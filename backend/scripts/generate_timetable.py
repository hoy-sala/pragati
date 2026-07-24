#!/usr/bin/env python3
"""
Master School Timetable Generator
Morarji Desai Residential School (SC-32), Bahaddurghatta, Chitradurga
Karnataka Residential Educational Institutions Society
Academic Year 2026-27

Usage: python generate_timetable.py [output_path]
"""

import sys, os
from collections import OrderedDict

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl is required. Install it with: pip install openpyxl")
    sys.exit(1)

# ─── COLOUR PALETTE ────────────────────────────────────────────────
COLOURS = {
    "KAN": "E0F2FE", "ENG": "E0E7FF", "HIN": "F3E8FF",
    "MAT": "FFEDD5", "SCI": "FEF08A", "SOC": "FED7AA",
    "CS":  "CFFAFE", "DRW": "FCE7F3", "MUS": "FEE2E2",
    "PE":  "D1FAE5", "LIB": "E2E8F0", "CUL": "DCFCE7",
    "BRK": "FEF3C7", "LNH": "FDE68A",
}

SUBJECT_NAMES = OrderedDict([
    ("KAN", "Kannada"),
    ("ENG", "English"),
    ("HIN", "Hindi"),
    ("MAT", "Mathematics"),
    ("SCI", "Science (Phy/Chem/Bio)"),
    ("SOC", "Social Studies"),
    ("CS",  "Computer Science"),
    ("DRW", "Drawing & Visual Arts"),
    ("MUS", "Music & Performing Arts"),
    ("PE",  "Physical Education"),
    ("LIB", "Library & Reading"),
    ("CUL", "Cultural Programme"),
])

# ─── WEEKLY ALLOCATION ─────────────────────────────────────────────
ALLOC = OrderedDict([
    ("KAN", 6), ("MAT", 6), ("SCI", 6), ("ENG", 5), ("SOC", 5),
    ("HIN", 4), ("CS", 2), ("DRW", 2), ("MUS", 2), ("PE", 2),
    ("LIB", 2), ("CUL", 2),
])
assert sum(ALLOC.values()) == 44, f"Total periods must be 44, got {sum(ALLOC.values())}"

# ─── DAYS AND PERIODS ──────────────────────────────────────────────
DAYS = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]
WEEKDAY_PERIODS = 8
SAT_PERIODS = 4

# ─── CONSTRUCT TIMETABLE ───────────────────────────────────────────
# 44-period schedules for Classes 6-10 satisfying all constraints.
# Each schedule is a dict[day][period_index] = subject_code.
# CUL is placed on Fri periods 6,7 (0-indexed: indices 6,7).

def build_rotation(base_variants):
    """Return 5 schedules (one per class) cycling through variants."""
    return [base_variants[i % len(base_variants)] for i in range(5)]

# I design 5 distinct valid schedules (one per class) that share the
# same P1 lock and CUL placement but vary in the other periods.
#
# Layout: each schedule has keys "Mon".."Fri" (list of 8 codes) and
#          "Sat" (list of 4 codes).
#
# Constraint verification is performed below.

SCHEDULE_CLASS6 = {
    "Mon": ["KAN","MAT","ENG","SCI","SOC","HIN","CS","DRW"],
    "Tue": ["MAT","SCI","KAN","ENG","HIN","SOC","PE","MUS"],
    "Wed": ["ENG","KAN","MAT","SOC","SCI","CS","LIB","HIN"],
    "Thu": ["SCI","SOC","ENG","KAN","MAT","DRW","MUS","PE"],
    "Fri": ["HIN","KAN","SCI","MAT","SOC","ENG","CUL","CUL"],
    "Sat": ["MAT","LIB","KAN","SCI"],
}

SCHEDULE_CLASS7 = {
    "Mon": ["MAT","KAN","SCI","ENG","SOC","PE","HIN","SOC"],
    "Tue": ["ENG","MAT","HIN","KAN","SCI","DRW","PE","CS"],
    "Wed": ["KAN","SCI","ENG","SOC","MAT","MUS","LIB","DRW"],
    "Thu": ["SOC","HIN","MAT","KAN","ENG","CS","MUS","SCI"],
    "Fri": ["SCI","ENG","KAN","MAT","SOC","HIN","CUL","CUL"],
    "Sat": ["KAN","LIB","MAT","SCI"],
}

SCHEDULE_CLASS8 = {
    "Mon": ["ENG","SCI","MAT","KAN","SOC","MUS","SOC","LIB"],
    "Tue": ["KAN","MAT","SCI","SOC","ENG","CS","DRW","PE"],
    "Wed": ["MAT","ENG","KAN","SCI","HIN","DRW","SOC","MUS"],
    "Thu": ["HIN","KAN","MAT","ENG","SCI","LIB","PE","CS"],
    "Fri": ["SCI","SOC","KAN","ENG","MAT","HIN","CUL","CUL"],
    "Sat": ["KAN","HIN","SCI","MAT"],
}

SCHEDULE_CLASS9 = {
    "Mon": ["SCI","SOC","ENG","KAN","MAT","LIB","HIN","SOC"],
    "Tue": ["MAT","SCI","ENG","HIN","KAN","MUS","CS","DRW"],
    "Wed": ["KAN","MAT","SOC","SCI","HIN","PE","DRW","LIB"],
    "Thu": ["MAT","KAN","MAT","SOC","ENG","CS","MUS","SCI"],
    "Fri": ["ENG","KAN","SCI","MAT","SOC","HIN","CUL","CUL"],
    "Sat": ["SCI","PE","KAN","ENG"],
}

SCHEDULE_CLASS10 = {
    "Mon": ["MAT","SCI","KAN","ENG","HIN","CS","SOC","DRW"],
    "Tue": ["SCI","MAT","ENG","KAN","SOC","LIB","MUS","PE"],
    "Wed": ["HIN","ENG","SOC","MAT","KAN","SCI","CS","MUS"],
    "Thu": ["ENG","KAN","HIN","SOC","MAT","PE","LIB","SCI"],
    "Fri": ["KAN","ENG","SCI","MAT","HIN","SOC","CUL","CUL"],
    "Sat": ["SCI","DRW","MAT","KAN"],
}

SCHEDULES = [SCHEDULE_CLASS6, SCHEDULE_CLASS7, SCHEDULE_CLASS8,
             SCHEDULE_CLASS9, SCHEDULE_CLASS10]
CLASS_NAMES = ["Class 6", "Class 7", "Class 8", "Class 9", "Class 10"]

# ─── CONSTRAINT VERIFICATION ───────────────────────────────────────
def verify_schedules():
    errors = []
    for ci, sched in enumerate(SCHEDULES):
        name = CLASS_NAMES[ci]
        counts = {}
        for day in DAYS:
            periods = sched[day]
            prev = None
            for pi, code in enumerate(periods):
                counts[code] = counts.get(code, 0) + 1
                # No consecutive duplicates (allow CUL-CUL on Fri)
                if code == prev and not (day == "Fri" and code == "CUL" and pi >= 6):
                    errors.append(f"{name} {day} P{pi+1}: consecutive {code}")
                prev = code
                # P1 lock: only core subjects
                if pi == 0 and code not in ("KAN","ENG","HIN","MAT","SCI","SOC"):
                    errors.append(f"{name} {day} P1: non-core {code}")

        for subj, expected in ALLOC.items():
            actual = counts.get(subj, 0)
            if actual != expected:
                errors.append(f"{name} {subj}: got {actual}, expected {expected}")

        # CUL on Fri P7,P8 (indices 6,7)
        if sched["Fri"][6] != "CUL" or sched["Fri"][7] != "CUL":
            errors.append(f"{name} Fri P7,P8: CUL missing")

    if errors:
        print("CONSTRAINT VIOLATIONS FOUND:")
        for e in errors:
            print(f"  - {e}")
        sys.exit(1)
    else:
        print("All constraints verified: 5 schedules, 220 periods, zero violations.")

verify_schedules()

# ─── EXCEL GENERATION ──────────────────────────────────────────────
DAYS_LABEL = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
DAY_ABBR  = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]

# Period start times
WD_START = ["10:00 AM","10:40 AM","11:20 AM","12:10 PM","12:50 PM",
            "02:20 PM","03:00 PM","03:40 PM"]
WD_END   = ["10:40 AM","11:20 AM","12:00 PM","12:50 PM","01:30 PM",
            "03:00 PM","03:40 PM","04:20 PM"]
SAT_START = ["09:50 AM","10:30 AM","11:10 AM","11:50 AM"]
SAT_END   = ["10:30 AM","11:10 AM","11:50 AM","12:30 PM"]

ROWS_BETWEEN_CLASSES = 2   # blank rows between class blocks
START_ROW = 7               # data starts at row 7
COL_CLASS = 1               # A = class name
COL_DAY   = 2               # B = day label
COL_TIME  = 3               # C = time range
COL_P1    = 4               # D..K = periods P1-P8 (weekday) / P1-P4 (sat)
COL_P2    = 5
COL_P3    = 6
COL_P4    = 7
COL_P5    = 8
COL_P6    = 9
COL_P7    = 10
COL_P8    = 11

# Saturday only uses COL_P1..COL_P4
SAT_COL_START = COL_P1
SAT_COL_END   = COL_P4

def fill_cell(ws, row, col, value, fill_color=None, font=None, alignment=None,
              border=None, number_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    if fill_color:
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    if font:
        cell.font = font
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    if number_format:
        cell.number_format = number_format
    return cell

def generate(output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Master Timetable 2026-27"

    # ─── Column Widths ─────────────────────────────────────────────
    ws.column_dimensions["A"].width = 11
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 20
    for c in range(4, 12):
        ws.column_dimensions[get_column_letter(c)].width = 14

    # ─── Styles ────────────────────────────────────────────────────
    header_font = Font(name="Trebuchet MS", size=16, bold=True, color="1E293B")
    sub_font = Font(name="Trebuchet MS", size=11, bold=False, color="475569")
    day_font = Font(name="Segoe UI", size=10, bold=True, color="1E293B")
    period_font = Font(name="Segoe UI", size=10, bold=False, color="1E293B")
    period_font_small = Font(name="Segoe UI", size=9, bold=False, color="334155")
    time_font = Font(name="Segoe UI", size=9, bold=False, color="64748B")
    legend_header_font = Font(name="Segoe UI", size=10, bold=True, color="1E293B")
    legend_font = Font(name="Segoe UI", size=9, bold=False, color="334155")

    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    thick_bottom = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="medium", color="475569"),
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    header_align = Alignment(horizontal="center", vertical="center")

    # ─── HEADER BLOCK ──────────────────────────────────────────────
    ws.merge_cells("A1:K1")
    fill_cell(ws, 1, 1, "KARNATAKA RESIDENTIAL EDUCATIONAL INSTITUTIONS SOCIETY",
              fill_color="1E293B", font=Font(name="Trebuchet MS", size=10, bold=False, color="94A3B8"),
              alignment=Alignment(horizontal="center"))
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:K2")
    fill_cell(ws, 2, 1, "MORARJI DESAI RESIDENTIAL SCHOOL (SC-32) BAHADDURGHATTA, CHITRADURGA",
              fill_color="1E293B", font=Font(name="Trebuchet MS", size=12, bold=True, color="F8FAFC"),
              alignment=Alignment(horizontal="center"))
    ws.row_dimensions[2].height = 26

    ws.merge_cells("A3:K3")
    fill_cell(ws, 3, 1, "MASTER SCHOOL TIME TABLE 2026-27",
              fill_color="2563EB", font=Font(name="Trebuchet MS", size=18, bold=True, color="FFFFFF"),
              alignment=Alignment(horizontal="center"))
    ws.row_dimensions[3].height = 38

    # Sub-header row
    ws.merge_cells("A4:K4")
    fill_cell(ws, 4, 1,
        "MONDAY – FRIDAY: 10:00 AM – 04:20 PM  |  SATURDAY: 09:50 AM – 12:30 PM  |  "
        "BREAK: 12:00–12:10 PM  |  LUNCH: 01:30–02:20 PM",
        fill_color="F1F5F9", font=sub_font, alignment=Alignment(horizontal="center"))
    ws.row_dimensions[4].height = 22

    # Blank spacer
    ws.row_dimensions[5].height = 6

    # Column headers (row 6)
    headers = ["Class", "Day", "Time", "P1", "P2", "P3", "P4", "P5", "P6", "P7", "P8"]
    for ci, h in enumerate(headers, 1):
        f = header_font if ci <= 3 else period_font
        fill_cell(ws, 6, ci, h, fill_color="E2E8F0", font=f,
                  alignment=header_align, border=thin_border)
    ws.row_dimensions[6].height = 24

    # ─── FREEZE PANES ──────────────────────────────────────────────
    ws.freeze_panes = "C7"

    # ─── TIMETABLE DATA ROWS ───────────────────────────────────────
    current_row = START_ROW

    # Break / Lunch labels
    BREAK_LABEL = "BRK"
    LUNCH_LABEL = "LNH"
    BREAK_ROW_LABEL = "Short Break"
    LUNCH_ROW_LABEL = "Lunch Break"
    BREAK_TIME = "12:00 PM – 12:10 PM"
    LUNCH_TIME = "01:30 PM – 02:20 PM"

    WEEKDAY_BREAK_INDEX = 3   # after P3 (0-indexed)
    WEEKDAY_LUNCH_INDEX = 5   # after P5 (0-indexed)

    for ci, sched in enumerate(SCHEDULES):
        class_name = CLASS_NAMES[ci]
        first_row_this_class = current_row

        for di, day_key in enumerate(DAYS):
            periods = sched[day_key]
            is_saturday = (day_key == "Sat")
            num_periods = SAT_PERIODS if is_saturday else WEEKDAY_PERIODS

            times = []
            if is_saturday:
                for pi in range(num_periods):
                    times.append(f"{SAT_START[pi]} – {SAT_END[pi]}")
            else:
                for pi in range(num_periods):
                    times.append(f"{WD_START[pi]} – {WD_END[pi]}")

            row = current_row

            # Day label (merged across period columns for the day)
            day_label = DAYS_LABEL[di]

            # Write Class name in first row only, merge downwards
            if di == 0:
                fill_cell(ws, row, COL_CLASS, class_name,
                          font=Font(name="Segoe UI", size=10, bold=True, color="1E293B"),
                          alignment=center_align)

            # Write day row
            fill_cell(ws, row, COL_DAY, day_label,
                      font=Font(name="Segoe UI", size=9, bold=True, color="475569"),
                      alignment=center_align)

            # Build period rows
            for pi, code in enumerate(periods):
                col = COL_P1 + pi
                is_core_subject = code in ("KAN","ENG","HIN","MAT","SCI","SOC")

                if code in COLOURS:
                    bg = COLOURS[code]
                else:
                    bg = "FFFFFF"

                # Determine font
                if code in (BREAK_LABEL, LUNCH_LABEL):
                    f = period_font_small
                elif is_core_subject:
                    f = period_font
                else:
                    f = period_font

                is_last_class_in_day = (ci == len(SCHEDULES) - 1)
                bdr = thick_bottom if is_last_class_in_day else thin_border

                fill_cell(ws, row, col, code, fill_color=bg, font=f,
                          alignment=center_align, border=bdr)

            # Merge columns for time
            time_val = "; ".join(times)
            fill_cell(ws, row, COL_TIME, time_val,
                      font=time_font, alignment=left_align, border=thin_border)

            # Fill empty period cells
            for col in range(COL_P1, COL_P8 + 1):
                if ws.cell(row=row, column=col).value is None:
                    fill_cell(ws, row, col, "", border=thin_border)

            # Add break/lunch annotation rows (insert extra rows)
            # Weekday: after P3 and P5
            if not is_saturday:
                # Short Break after P3
                br_row = row
                col = COL_P1 + WEEKDAY_BREAK_INDEX
                if periods[WEEKDAY_BREAK_INDEX] != BREAK_LABEL:
                    # Insert a row for break label below this period row
                    pass  # breaks are already coded as BRK in the time slot row

            # Day separator line handled by thick bottom border on class 10

            current_row += 1

        # Add blank row between classes
        current_row += 1

    # ─── SUBJECT LEGEND ────────────────────────────────────────────
    current_row += 1
    fill_cell(ws, current_row, 1, "SUBJECT LEGEND",
              font=Font(name="Segoe UI", size=12, bold=True, color="1E293B"))
    ws.merge_cells(start_row=current_row, start_column=1,
                   end_row=current_row, end_column=6)
    current_row += 1

    legend_items = list(SUBJECT_NAMES.items())
    # Two-column layout for legend
    mid = (len(legend_items) + 1) // 2
    for i, (code, name) in enumerate(legend_items):
        col = 1 if i < mid else 4
        row_offset = i if i < mid else i - mid
        r = current_row + row_offset
        bg = COLOURS.get(code, "FFFFFF")
        fill_cell(ws, r, col, f"{code} = {name}",
                  fill_color=bg, font=legend_font, alignment=left_align,
                  border=thin_border)
        ws.merge_cells(start_row=r, start_column=col,
                       end_row=r, end_column=col + 2)
    current_row += len(legend_items[:mid]) + 1

    # ─── GENERATION INFO ───────────────────────────────────────────
    current_row += 1
    fill_cell(ws, current_row, 1,
        "Generated by KREIS Master Schedule Architect | Academic Year 2026-27",
        font=Font(name="Segoe UI", size=8, italic=True, color="94A3B8"))
    ws.merge_cells(start_row=current_row, start_column=1,
                   end_row=current_row, end_column=11)

    # ─── SAVE ──────────────────────────────────────────────────────
    wb.save(output_path)
    print(f"Timetable saved to: {output_path}")
    print(f"File size: {os.path.getsize(output_path):,} bytes")


if __name__ == "__main__":
    out = sys.argv[1] if len(sys.argv) > 1 else \
        os.path.join(os.path.dirname(__file__), "..", "..",
                     "frontend", "static", "timetable",
                     "Master_Timetable_2026-27.xlsx")
    out = os.path.abspath(out)
    os.makedirs(os.path.dirname(out), exist_ok=True)
    generate(out)
